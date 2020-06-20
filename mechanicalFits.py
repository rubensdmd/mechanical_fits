#! python3
# mechanicalFits.py - A mechanical fits reader, given a basic size, 
# a fundamental deviation and a international tolerance grade as arguments.

from pathlib import Path
import openpyxl, re

def tolerance(diametro,pos,calidad):

    # Abrir documento de Excel
    wb = openpyxl.load_workbook('Tolerancias.xlsx')

    # Ingresar diámetro nominal
    if diametro < 0 or diametro >= 500:
        return 'The basic size argument is out of range'

    # Ingresar posición:
    Ejes = ['a', 'b', 'c', 'cd', 'd', 'e', 'ef', 'f', 'fg', 'g', 'h', 'j', 'js', 'k', 'm', 'n', 'p', 'r', 's', 't', 'u', 'v', 'x', 'y', 'z', 'za', 'zb', 'zc']
    Agujeros = ['A', 'B', 'C', 'CD', 'D', 'E', 'EF', 'F', 'FG', 'G', 'H', 'J', 'JS', 'K', 'M', 'N', 'P', 'R', 'S', 'T', 'U', 'V', 'X', 'Y', 'Z', 'ZA', 'ZB', 'ZC']

    if diametro <= 24:
        Ejes.remove('t')
        Agujeros.remove('T') 
    elif diametro <= 18:
        Ejes.remove('y')
        Agujeros.remove('Y')
    elif diametro <= 14:
        Ejes.remove('v')
        Agujeros.remove('V')

    if diametro > 10:
        Ejes = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'j', 'js', 'k', 'm', 'n', 'p', 'r', 's', 't', 'u', 'v', 'x', 'y', 'z', 'za', 'zb', 'zc']
        Agujeros = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'JS', 'K', 'M', 'N', 'P', 'R', 'S', 'T', 'U', 'V', 'X', 'Y', 'Z', 'ZA', 'ZB', 'ZC']

    Bases = []
    for i in range(len(Ejes)): 
        Bases.append(Ejes[i])
    for i in range(len(Agujeros)):
        Bases.append(Agujeros[i]) 

    if pos not in Bases:
        return 'The fundamental deviation argument is not valid'

    # Selección de la hoja de cálculo a usar

    if pos in Ejes:
        sheet = wb['Eje']
    else:
        sheet = wb['Agujero']

    # Ingresar calidad
    
    calidades = ['0','01','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18']

    if pos == 'j':
        calidades = ['5','6','7','8']
    if pos == 'J':
        calidades = ['6','7','8']

    #calidad = pyip.inputChoice(calidades,prompt="Grado de la tolerancia: ")
    if calidad not in calidades:
        return 'The international tolerance grade provide is not valid'
    
    ISOcalidad = 'IT '+ calidad

    # Lectura del valor de la calidad: gradoTol
    gradoSheet = wb['Grado tolerancia']
    diamRegex = re.compile(r'\d+')

    for row in range(3, gradoSheet.max_row + 1):
        grado = diamRegex.findall(gradoSheet['A' + str(row)].value)
        if int(grado[0]) < diametro <= int(grado [1]):
            fila = str(row)
            break

    from openpyxl.utils import get_column_letter, column_index_from_string
    for column in range(2, gradoSheet.max_column + 1):
        if ISOcalidad == gradoSheet[get_column_letter(column) + '2'].value:
            columna = get_column_letter(column)
            break

    gradoTol = gradoSheet[columna + fila].value

    # Lectura de la desviación

    if pos == 'js' or pos == 'JS':  # Caso particular (tolerancia simétrica)
        tolerancia = gradoTol/2
        print('Banda de tolerancia:\n'+ '+' + str(tolerancia) + '\n' + '-' + str(tolerancia))
        exit()

    for row in range(4, sheet.max_row + 1):
        grado = diamRegex.findall(sheet['A' + str(row)].value)
        if int(grado[0]) < diametro <= int(grado [1]):
            fila = str(row)
            break

    for column in range(2, sheet.max_column + 1):
        if pos == sheet[get_column_letter(column) + '2'].value:
            columna = get_column_letter(column)
            break

    if pos == 'j':    #Caso particular
        if calidad == '7':
            columna = get_column_letter(column_index_from_string(columna) + 1) 
        if calidad == '8':
            columna = get_column_letter(column_index_from_string(columna) + 2) 
            if diametro > 3:
                print('El ajuste requerido no exite, vuelva a intentarlo')
                exit()

    if pos == 'J':    #Caso particular
        if calidad == '7':
            columna = get_column_letter(column_index_from_string(columna) + 1) 
        if calidad == '8':
            columna = get_column_letter(column_index_from_string(columna) + 2) 

    if pos == 'k':  #Caso particular
        if 4 >= int(calidad) or int(calidad) >= 7:
            columna = get_column_letter(column_index_from_string(columna) + 1)

    if pos == 'K':  #Caso particular
        if int(calidad) > 8:
            columna = get_column_letter(column_index_from_string(columna) + 1)    
            if diametro > 3:
                print('El ajuste requerido no exite, vuelva a intentarlo')
                exit()

    # Cálculo banda de tolerancia 

    tolerancia = sheet[columna + fila].value

    if pos == 'K' or pos == 'M':
        if int(calidad) <= 8 and diametro > 3:
            for column in range(int(column_index_from_string('AJ')), int(column_index_from_string('AO'))+1):
                numRegex = re.compile(r'\d+')
                tole = numRegex.findall(tolerancia)
                if ISOcalidad == sheet[get_column_letter(column) + '3'].value:
                    columna = get_column_letter(column)
                    break
            delta = sheet[columna + fila].value
            tolerancia = -int(tole[0]) + delta
        else:
            columna = get_column_letter(column_index_from_string(columna) + 1)  

    if pos == 'N':
        if int(calidad) <= 8:
            for column in range(int(column_index_from_string('AJ')), int(column_index_from_string('AO'))+1):
                numRegex = re.compile(r'\d+')
                tole = numRegex.findall(tolerancia)
                if ISOcalidad == sheet[get_column_letter(column) + '3'].value:
                    columna = get_column_letter(column)
                    break
            delta = sheet[columna + fila].value
            tolerancia = -int(tole[0]) + delta
        else:
            columna = get_column_letter(column_index_from_string(columna) + 1)  

    if tolerancia == None:  # Ajuste para celdas combinadas
        tolerancia = sheet[columna + str(int(fila)-1)].value

    PaZC = ['P', 'R', 'S', 'T', 'U', 'V', 'X', 'Y', 'Z', 'ZA', 'ZB', 'ZC']

    if pos in PaZC:
        if int(calidad) <= 7:
            for column in range(int(column_index_from_string('AJ')), int(column_index_from_string('AO'))+1):
                numRegex = re.compile(r'\d+')
                tole = numRegex.findall(tolerancia)
                if ISOcalidad == sheet[get_column_letter(column) + '3'].value:
                    columna = get_column_letter(column)
                    break
            delta = sheet[columna + fila].value
            tolerancia = -int(tole[0]) + delta

    JaZC = ['J', 'K', 'M', 'N', 'P', 'R', 'S', 'T', 'U', 'V', 'X', 'Y', 'Z', 'ZA', 'ZB', 'ZC']

    if pos in Ejes:
        print('Banda de tolerancia:\n'+ str(tolerancia) + '\n' + str(tolerancia - gradoTol))
    else:
        if pos in JaZC:
            print('Banda de tolerancia:\n'+ str(tolerancia) + '\n' + str(tolerancia - gradoTol))
        else:
            print('Banda de tolerancia:\n' + str(tolerancia + gradoTol) + '\n' + str(tolerancia))

prueba = tolerance(50,'H','8')


